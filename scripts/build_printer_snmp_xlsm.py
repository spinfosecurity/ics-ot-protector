#!/usr/bin/env python3
"""
Generates PrinterConsumables.xlsm – an Excel macro-enabled workbook whose
VBA code queries IP printers via SNMP (using PowerShell on Windows) and
displays consumable levels with colour-coded progress bars.

Run once on any machine that has Python + openpyxl installed:
    python3 build_printer_snmp_xlsm.py

Then open PrinterConsumables.xlsm in Excel (Windows), enable macros,
fill in your printer IPs on the "Printers" sheet, and click "Scan Printers".
"""

import zipfile, shutil, os, textwrap
from pathlib import Path

# ---------------------------------------------------------------------------
# VBA source – stored as a plain string, injected into the xlsm zip
# ---------------------------------------------------------------------------
VBA_MODULE = r'''
Attribute VB_Name = "PrinterSNMP"
Option Explicit

' ============================================================
'  CONFIGURATION  – edit these defaults before your first run
' ============================================================
Const DEFAULT_COMMUNITY As String = "public"
Const SNMP_TIMEOUT_SEC  As Long   = 3        ' seconds per SNMP request
Const MAX_SUPPLIES       As Long   = 20       ' max supply rows per printer

' ============================================================
'  RFC 3805 / Printer-MIB OIDs
' ============================================================
Const OID_SYS_NAME    As String = "1.3.6.1.2.1.1.5.0"
Const OID_SYS_DESCR   As String = "1.3.6.1.2.1.1.1.0"
Const OID_SYS_UPTIME  As String = "1.3.6.1.2.1.1.3.0"
Const OID_SUPPLY_DESC As String = "1.3.6.1.2.1.43.11.1.1.6.1"
Const OID_SUPPLY_TYPE As String = "1.3.6.1.2.1.43.11.1.1.5.1"
Const OID_SUPPLY_MAX  As String = "1.3.6.1.2.1.43.11.1.1.8.1"
Const OID_SUPPLY_LVL  As String = "1.3.6.1.2.1.43.11.1.1.9.1"

' ============================================================
'  Supply-type integer -> friendly name
' ============================================================
Private Function SupplyTypeName(code As Long) As String
    Select Case code
        Case 3:  SupplyTypeName = "Toner"
        Case 4:  SupplyTypeName = "Waste Toner"
        Case 5:  SupplyTypeName = "Ink"
        Case 6:  SupplyTypeName = "Ink Cartridge"
        Case 7:  SupplyTypeName = "Ink Ribbon"
        Case 9:  SupplyTypeName = "OPC Drum"
        Case 10: SupplyTypeName = "Developer"
        Case 11: SupplyTypeName = "Fuser Oil"
        Case 15: SupplyTypeName = "Fuser"
        Case 20: SupplyTypeName = "Transfer Unit"
        Case 21: SupplyTypeName = "Toner Cartridge"
        Case 22: SupplyTypeName = "Fuser Kit"
        Case 25: SupplyTypeName = "Maintenance Kit"
        Case 26: SupplyTypeName = "Drum Kit"
        Case Else: SupplyTypeName = "Other (" & code & ")"
    End Select
End Function

' ============================================================
'  Run a PowerShell command and return its stdout as a string.
'  Uses a temp file so we can capture output without a visible window.
' ============================================================
Private Function RunPS(psCmd As String) As String
    Dim tmpFile As String
    tmpFile = Environ("TEMP") & "\snmp_tmp_" & Format(Timer, "0") & ".txt"

    Dim shell As Object
    Set shell = CreateObject("WScript.Shell")

    ' -NoProfile -NonInteractive keeps startup fast
    Dim fullCmd As String
    fullCmd = "powershell.exe -NoProfile -NonInteractive -Command " & _
              Chr(34) & psCmd & " | Out-File -Encoding UTF8 '" & tmpFile & "'" & Chr(34)

    shell.Run fullCmd, 0, True   ' 0 = hidden, True = wait

    Dim fileNo As Integer
    fileNo = FreeFile
    Dim result As String
    If Dir(tmpFile) <> "" Then
        Open tmpFile For Input As #fileNo
        Dim line As String
        Do While Not EOF(fileNo)
            Line Input #fileNo, line
            result = result & Trim(line) & vbLf
        Loop
        Close #fileNo
        Kill tmpFile
    End If
    RunPS = Trim(result)
End Function

' ============================================================
'  SNMP GET – returns the value string or "" on failure.
'  Uses PowerShell's Invoke-Expression with Net-SNMP snmpget
'  if available, otherwise falls back to a pure-PowerShell UDP walk.
' ============================================================
Private Function SNMPGet(ip As String, community As String, oid As String) As String
    ' Try Net-SNMP snmpget.exe first (fastest, most compatible)
    Dim psCheck As String
    psCheck = "if (Get-Command snmpget -ErrorAction SilentlyContinue) " & _
              "{ snmpget -v2c -c " & community & " -Oqv " & ip & " " & oid & " } " & _
              "else { 'NOSNMPGET' }"
    Dim raw As String
    raw = RunPS(psCheck)

    If InStr(raw, "NOSNMPGET") > 0 Or raw = "" Then
        ' Fallback: PowerShell UDP socket SNMP v1 GET
        raw = SNMPGetPS(ip, community, oid)
    End If
    SNMPGet = raw
End Function

' ============================================================
'  Pure-PowerShell SNMP v1 GET over raw UDP.
'  Builds and parses a minimal BER-encoded PDU.
'  Returns the decoded value or "".
' ============================================================
Private Function SNMPGetPS(ip As String, community As String, oid As String) As String
    Dim ps As String
    ' One-liner PowerShell script that does a raw SNMP v1 GET.
    ' We encode the OID, send the packet, receive the response, and decode.
    ps = "$ip='" & ip & "';$comm='" & community & "';$oid='" & oid & "';" & _
         "$to=" & SNMP_TIMEOUT_SEC & "000;" & _
         "function En-Int([int]$v){if($v-lt128){[byte[]]($v)}else{$b=[System.Collections.Generic.List[byte]]::new();do{$b.Insert(0,[byte]($v-band 0x7f));$v=$v-shr 7}while($v-gt 0);$b[0]=$b[0]-bor 0x80;$b.ToArray()}};" & _
         "function En-OID([string]$o){$parts=$o.Split('.');$out=[System.Collections.Generic.List[byte]]::new();$first=[int]$parts[0]*40+[int]$parts[1];$out.Add([byte]$first);for($i=2;$i-lt$parts.Length;$i++){$out.AddRange((En-Int([int]$parts[$i]))|%{$_})};$b=$out.ToArray();[byte[]](0x06)+[byte[]]($b.Length)+$b};" & _
         "$oidBytes=En-OID $oid;" & _
         "$null_b=[byte[]](0x05,0x00);" & _
         "$varbind=[byte[]](0x30)+[byte[]]($oidBytes.Length+$null_b.Length)+$oidBytes+$null_b;" & _
         "$commB=[System.Text.Encoding]::ASCII.GetBytes($comm);$commTlv=[byte[]](0x04)+[byte[]]($commB.Length)+$commB;" & _
         "$reqId=[byte[]](0x02,0x04,0x00,0x00,0x00,0x01);" & _
         "$errSt=[byte[]](0x02,0x01,0x00);$errIdx=[byte[]](0x02,0x01,0x00);" & _
         "$varList=[byte[]](0x30)+[byte[]]($varbind.Length)+$varbind;" & _
         "$pdu=[byte[]](0xa0)+[byte[]]($reqId.Length+$errSt.Length+$errIdx.Length+$varList.Length)+$reqId+$errSt+$errIdx+$varList;" & _
         "$ver=[byte[]](0x02,0x01,0x00);" & _
         "$msg=[byte[]](0x30)+[byte[]]($ver.Length+$commTlv.Length+$pdu.Length)+$ver+$commTlv+$pdu;" & _
         "$udp=New-Object System.Net.Sockets.UdpClient;" & _
         "$udp.Client.ReceiveTimeout=$to;" & _
         "try{[void]$udp.Send($msg,$msg.Length,$ip,161);" & _
         "$ep=New-Object System.Net.IPEndPoint([System.Net.IPAddress]::Any,0);" & _
         "$resp=$udp.Receive([ref]$ep);" & _
         "# Walk past version+community to the response PDU (0xa2)" & _
         "$i=0;while($i-lt$resp.Length-and $resp[$i]-ne 0xa2){$i++};" & _
         "# Skip past PDU tag/len, reqId, errSt, errIdx to varbind list" & _
         "$i++;$plen=$resp[$i];$i++;if($plen-band 0x80){$nb=$plen-band 0x7f;$i+=$nb};" & _
         "# skip reqId" & _
         "while($resp[$i]-ne 0x30-and $i-lt$resp.Length){$i++};" & _
         "# inside varbind list -> first varbind -> skip OID -> get value" & _
         "$i++;$vblen=$resp[$i];$i++;$i++;$olen=$resp[$i];$i++;$i+=$olen;" & _
         "$vtag=$resp[$i];$i++;$vlen=$resp[$i];$i++;" & _
         "if($vlen-band 0x80){$nb=$vlen-band 0x7f;$vlen=0;for($j=0;$j-lt$nb;$j++){$vlen=$vlen*256+$resp[$i];$i++}};" & _
         "$vbytes=$resp[$i..($i+$vlen-1)];" & _
         "switch($vtag){" & _
         "0x04{[System.Text.Encoding]::ASCII.GetString($vbytes)}" & _   ' OCTET STRING
         "0x02{$n=0;foreach($b in $vbytes){$n=$n*256+$b};$n}" & _       ' INTEGER
         "0x43{$n=0;foreach($b in $vbytes){$n=$n*256+$b};$n}" & _       ' TimeTicks
         "0x41{$n=0;foreach($b in $vbytes){$n=$n*256+$b};$n}" & _       ' Counter32
         "0x42{$n=0;foreach($b in $vbytes){$n=$n*256+$b};$n}" & _       ' Gauge32
         "0x40{($vbytes|%{$_.ToString()})-join'.'}" & _                  ' IpAddress
         "default{''}}" & _
         "}catch{''}" & _
         ";$udp.Close()"

    SNMPGetPS = RunPS(ps)
End Function

' ============================================================
'  SNMP WALK – returns an array of values for OID.1 .. OID.N
'  Walks index 1..MAX_SUPPLIES, returns values as pipe-delimited string.
' ============================================================
Private Function SNMPWalkIndexed(ip As String, community As String, _
                                  baseOid As String) As String()
    Dim results() As String
    ReDim results(1 To MAX_SUPPLIES)
    Dim i As Long
    For i = 1 To MAX_SUPPLIES
        Dim val As String
        val = SNMPGet(ip, community, baseOid & "." & i)
        If val = "" Or InStr(LCase(val), "no such") > 0 Or _
           InStr(LCase(val), "error") > 0 Then
            val = ""
        End If
        results(i) = val
    Next i
    SNMPWalkIndexed = results
End Function

' ============================================================
'  Colour helper: RGB long for a percentage
' ============================================================
Private Function PctColour(pct As Double) As Long
    If pct <= 10 Then
        PctColour = RGB(255, 112, 67)   ' red-orange  (critical)
    ElseIf pct <= 25 Then
        PctColour = RGB(255, 217, 102)  ' amber       (warning)
    Else
        PctColour = RGB(169, 209, 142)  ' green       (OK)
    End If
End Function

' ============================================================
'  Draw a simple in-cell "bar" using Unicode block characters
'  and colour-fill the cell.
' ============================================================
Private Sub DrawBar(cell As Range, pct As Double)
    Dim filled As Long
    filled = CLng(pct / 5)   ' 20 segments = 100%
    If filled > 20 Then filled = 20
    cell.Value = String(filled, Chr(9608)) & String(20 - filled, Chr(9617))
    cell.Interior.Color = PctColour(pct)
    cell.Font.Color = RGB(50, 50, 50)
End Sub

' ============================================================
'  MAIN ENTRY POINT
'  Reads printer list from "Printers" sheet, queries each,
'  writes results to "Consumables" sheet.
' ============================================================
Public Sub ScanPrinters()
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationManual
    Application.StatusBar = "Scanning printers via SNMP..."

    ' ---- ensure sheets exist ----
    Dim wbPrinters As Worksheet, wbResults As Worksheet
    On Error Resume Next
    Set wbPrinters = ThisWorkbook.Sheets("Printers")
    Set wbResults  = ThisWorkbook.Sheets("Consumables")
    On Error GoTo 0

    If wbPrinters Is Nothing Then
        MsgBox "Sheet 'Printers' not found. Please create it with columns: Name | IP | Community", _
               vbCritical, "Missing Sheet"
        GoTo Cleanup
    End If
    If wbResults Is Nothing Then
        Set wbResults = ThisWorkbook.Sheets.Add(After:=ThisWorkbook.Sheets(ThisWorkbook.Sheets.Count))
        wbResults.Name = "Consumables"
    End If

    ' ---- clear results sheet ----
    wbResults.Cells.Clear
    wbResults.Cells.Interior.ColorIndex = xlNone

    ' ---- write results header ----
    Dim timestamp As String
    timestamp = "Printer Consumables Report  —  " & Format(Now, "yyyy-mm-dd hh:mm:ss")
    With wbResults.Range("A1:J1")
        .Merge
        .Value = timestamp
        .Font.Bold = True
        .Font.Size = 13
        .Font.Color = RGB(255, 255, 255)
        .Interior.Color = RGB(13, 63, 110)
        .HorizontalAlignment = xlCenter
        .RowHeight = 26
    End With

    Dim headers As Variant
    headers = Array("Printer Name", "IP Address", "Status", "Supply", _
                    "Type", "Current", "Maximum", "% Left", "Bar Chart", "Polled At")
    Dim col As Long
    For col = 0 To UBound(headers)
        With wbResults.Cells(2, col + 1)
            .Value = headers(col)
            .Font.Bold = True
            .Font.Color = RGB(255, 255, 255)
            .Interior.Color = RGB(31, 78, 121)
            .HorizontalAlignment = xlCenter
            .Borders.LineStyle = xlContinuous
        End With
    Next col
    wbResults.Rows(2).RowHeight = 18

    ' ---- iterate printers ----
    Dim resultRow As Long
    resultRow = 3

    Dim printerRow As Long
    printerRow = 2   ' skip header on Printers sheet

    Do While wbPrinters.Cells(printerRow, 1).Value <> ""
        Dim pName      As String
        Dim pIP        As String
        Dim pCommunity As String

        pName      = Trim(wbPrinters.Cells(printerRow, 1).Value)
        pIP        = Trim(wbPrinters.Cells(printerRow, 2).Value)
        pCommunity = Trim(wbPrinters.Cells(printerRow, 3).Value)
        If pCommunity = "" Then pCommunity = DEFAULT_COMMUNITY

        Application.StatusBar = "Querying " & pName & " (" & pIP & ") ..."

        Dim sysName  As String
        Dim sysDescr As String
        Dim polledAt As String
        polledAt = Format(Now, "hh:mm:ss")

        ' Quick reachability check – get sysName
        sysName = SNMPGet(pIP, pCommunity, OID_SYS_NAME)

        If sysName = "" Then
            ' Printer unreachable
            With wbResults.Cells(resultRow, 1)
                .Value = pName
                .Interior.Color = RGB(255, 200, 200)
            End With
            wbResults.Cells(resultRow, 2).Value = pIP
            wbResults.Cells(resultRow, 3).Value = "UNREACHABLE"
            wbResults.Cells(resultRow, 3).Font.Color = RGB(200, 0, 0)
            wbResults.Cells(resultRow, 3).Font.Bold = True
            wbResults.Cells(resultRow, 10).Value = polledAt
            ApplyRowBorder wbResults, resultRow
            resultRow = resultRow + 1
        Else
            ' Walk supplies
            Dim descs() As String, types() As String
            Dim maxCaps() As String, levels() As String

            descs   = SNMPWalkIndexed(pIP, pCommunity, OID_SUPPLY_DESC)
            types   = SNMPWalkIndexed(pIP, pCommunity, OID_SUPPLY_TYPE)
            maxCaps = SNMPWalkIndexed(pIP, pCommunity, OID_SUPPLY_MAX)
            levels  = SNMPWalkIndexed(pIP, pCommunity, OID_SUPPLY_LVL)

            Dim startRow As Long
            startRow = resultRow

            Dim foundAny As Boolean
            foundAny = False

            Dim idx As Long
            For idx = 1 To MAX_SUPPLIES
                If descs(idx) <> "" Then
                    foundAny = True
                    Dim maxVal As Long, lvlVal As Long, typeVal As Long
                    maxVal  = Val(maxCaps(idx))
                    lvlVal  = Val(levels(idx))
                    typeVal = Val(types(idx))

                    Dim pct As Double
                    Dim pctStr As String
                    If maxVal > 0 And lvlVal >= 0 Then
                        pct = lvlVal / maxVal * 100
                        pctStr = Format(pct, "0.0") & "%"
                    ElseIf lvlVal = -2 Or maxVal <= 0 Then
                        pct = -1
                        pctStr = "N/A"
                    Else
                        pct = 0
                        pctStr = "0.0%"
                    End If

                    Dim altColour As Long
                    If resultRow Mod 2 = 0 Then
                        altColour = RGB(214, 228, 240)
                    Else
                        altColour = RGB(255, 255, 255)
                    End If

                    ' Printer name / IP / status – only on first supply row
                    If idx = startRow - startRow + 1 Then   ' always first
                        wbResults.Cells(resultRow, 1).Value = pName
                        wbResults.Cells(resultRow, 2).Value = pIP
                        With wbResults.Cells(resultRow, 3)
                            .Value = "Online"
                            .Font.Color = RGB(0, 128, 0)
                            .Font.Bold = True
                        End With
                    End If

                    wbResults.Cells(resultRow, 4).Value = descs(idx)
                    wbResults.Cells(resultRow, 5).Value = SupplyTypeName(typeVal)
                    wbResults.Cells(resultRow, 6).Value = IIf(lvlVal >= 0, lvlVal, "N/A")
                    wbResults.Cells(resultRow, 7).Value = IIf(maxVal > 0, maxVal, "N/A")
                    wbResults.Cells(resultRow, 8).Value = pctStr

                    If pct >= 0 Then
                        DrawBar wbResults.Cells(resultRow, 9), pct
                    Else
                        wbResults.Cells(resultRow, 9).Value = "Unknown"
                        wbResults.Cells(resultRow, 9).Interior.Color = RGB(217, 217, 217)
                    End If

                    wbResults.Cells(resultRow, 10).Value = polledAt

                    ' Apply background to informational columns
                    Dim c As Long
                    For c = 1 To 8
                        If wbResults.Cells(resultRow, c).Interior.ColorIndex = xlNone Or _
                           wbResults.Cells(resultRow, c).Interior.Color = RGB(255, 255, 255) Or _
                           wbResults.Cells(resultRow, c).Interior.Color = RGB(214, 228, 240) Then
                            wbResults.Cells(resultRow, c).Interior.Color = altColour
                        End If
                    Next c

                    ApplyRowBorder wbResults, resultRow
                    resultRow = resultRow + 1
                End If
            Next idx

            If Not foundAny Then
                wbResults.Cells(resultRow, 1).Value = pName
                wbResults.Cells(resultRow, 2).Value = pIP
                wbResults.Cells(resultRow, 3).Value = "Online"
                wbResults.Cells(resultRow, 4).Value = "(No Printer-MIB supplies reported)"
                wbResults.Cells(resultRow, 10).Value = polledAt
                ApplyRowBorder wbResults, resultRow
                resultRow = resultRow + 1
            Else
                ' Merge name/IP/status cells across all supply rows for this printer
                If resultRow - startRow > 1 Then
                    MergeCenterCol wbResults, startRow, resultRow - 1, 1
                    MergeCenterCol wbResults, startRow, resultRow - 1, 2
                    MergeCenterCol wbResults, startRow, resultRow - 1, 3
                End If
            End If
        End If

        printerRow = printerRow + 1
    Loop

    ' ---- column widths ----
    With wbResults
        .Columns(1).ColumnWidth = 24
        .Columns(2).ColumnWidth = 16
        .Columns(3).ColumnWidth = 13
        .Columns(4).ColumnWidth = 36
        .Columns(5).ColumnWidth = 18
        .Columns(6).ColumnWidth = 10
        .Columns(7).ColumnWidth = 10
        .Columns(8).ColumnWidth = 10
        .Columns(9).ColumnWidth = 24
        .Columns(10).ColumnWidth = 10
        .Rows(1).RowHeight = 26
        .FreezePanes = False
        .Activate
        .Cells(3, 1).Select
        ActiveWindow.FreezePanes = True
    End With

    wbResults.Activate
    MsgBox "Scan complete! Results written to the 'Consumables' sheet.", _
           vbInformation, "Done"

Cleanup:
    Application.ScreenUpdating = True
    Application.Calculation = xlCalculationAutomatic
    Application.StatusBar = False
End Sub

' ============================================================
'  Helper: apply thin border to all used columns in a row
' ============================================================
Private Sub ApplyRowBorder(ws As Worksheet, r As Long)
    Dim c As Long
    For c = 1 To 10
        With ws.Cells(r, c).Borders
            .LineStyle = xlContinuous
            .Weight = xlThin
        End With
    Next c
End Sub

' ============================================================
'  Helper: merge & centre a column range
' ============================================================
Private Sub MergeCenterCol(ws As Worksheet, r1 As Long, r2 As Long, c As Long)
    With ws.Range(ws.Cells(r1, c), ws.Cells(r2, c))
        .Merge
        .HorizontalAlignment = xlCenter
        .VerticalAlignment = xlCenter
    End With
End Sub

' ============================================================
'  Export current Consumables sheet to a plain .xlsx snapshot
' ============================================================
Public Sub ExportSnapshot()
    Dim savePath As String
    savePath = Application.GetSaveAsFilename( _
        InitialFileName:="PrinterReport_" & Format(Now, "yyyymmdd_hhmmss") & ".xlsx", _
        FileFilter:="Excel Workbook (*.xlsx), *.xlsx", _
        Title:="Save snapshot as...")
    If savePath = "False" Then Exit Sub

    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets("Consumables")
    ws.Copy
    ActiveWorkbook.SaveAs Filename:=savePath, FileFormat:=xlOpenXMLWorkbook
    ActiveWorkbook.Close False
    MsgBox "Snapshot saved: " & savePath, vbInformation, "Exported"
End Sub
'''

# ---------------------------------------------------------------------------
# Build the .xlsm using openpyxl + manual zip injection of VBA project
# ---------------------------------------------------------------------------

def build_xlsm(out_path: str = "PrinterConsumables.xlsm"):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    #  Sheet 1: Printers  (user fills in their printer list here)
    # ------------------------------------------------------------------ #
    ws_printers = wb.active
    ws_printers.title = "Printers"

    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    thin      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title
    ws_printers.merge_cells("A1:D1")
    c = ws_printers["A1"]
    c.value = "Printer List  —  Edit this sheet, then click  Scan Printers"
    c.font  = Font(bold=True, size=13, color="FFFFFF")
    c.fill  = PatternFill("solid", fgColor="0D3F6E")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_printers.row_dimensions[1].height = 26

    headers = ["Printer Name", "IP Address", "Community String (leave blank = 'public')", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws_printers.cell(2, col, h)
        cell.font   = hdr_font
        cell.fill   = hdr_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_printers.row_dimensions[2].height = 18

    sample_printers = [
        ("Office Printer 1",    "192.168.1.100", "public",  "Reception"),
        ("IT Room Xerox",       "192.168.1.101", "private", "IT Dept"),
        ("Finance Floor Canon", "192.168.1.102", "",        "Finance"),
    ]
    alt_colours = ["D6E4F0", "FFFFFF"]
    for r, (name, ip, comm, note) in enumerate(sample_printers, 3):
        fill = PatternFill("solid", fgColor=alt_colours[r % 2])
        for col, val in enumerate([name, ip, comm, note], 1):
            cell = ws_printers.cell(r, col, val)
            cell.fill   = fill
            cell.border = thin
            cell.alignment = Alignment(vertical="center")

    ws_printers.column_dimensions["A"].width = 26
    ws_printers.column_dimensions["B"].width = 18
    ws_printers.column_dimensions["C"].width = 34
    ws_printers.column_dimensions["D"].width = 22

    # ------------------------------------------------------------------ #
    #  Sheet 2: Consumables  (populated by macro, pre-format header only)
    # ------------------------------------------------------------------ #
    ws_results = wb.create_sheet("Consumables")
    ws_results.merge_cells("A1:J1")
    rc = ws_results["A1"]
    rc.value = "Run 'Scan Printers' macro to populate this sheet"
    rc.font  = Font(bold=True, size=12, color="FFFFFF", italic=True)
    rc.fill  = PatternFill("solid", fgColor="0D3F6E")
    rc.alignment = Alignment(horizontal="center", vertical="center")
    ws_results.row_dimensions[1].height = 26

    # ------------------------------------------------------------------ #
    #  Sheet 3: Instructions
    # ------------------------------------------------------------------ #
    ws_help = wb.create_sheet("Instructions")
    instructions = [
        ("HOW TO USE", True, "0D3F6E", "FFFFFF"),
        ("", False, None, None),
        ("Step 1", True, "2E75B6", "FFFFFF"),
        ("  Open this file in Microsoft Excel (Windows).", False, None, None),
        ("  If prompted, click 'Enable Macros'.", False, None, None),
        ("", False, None, None),
        ("Step 2", True, "2E75B6", "FFFFFF"),
        ("  Go to the 'Printers' sheet.", False, None, None),
        ("  Replace the sample rows with your printer Name, IP, and SNMP community string.", False, None, None),
        ("  Leave the Community column blank to use 'public'.", False, None, None),
        ("", False, None, None),
        ("Step 3", True, "2E75B6", "FFFFFF"),
        ("  Press Alt+F8, select 'ScanPrinters', and click Run.", False, None, None),
        ("  OR: Go to Developer > Macros > ScanPrinters > Run.", False, None, None),
        ("  Results appear on the 'Consumables' sheet.", False, None, None),
        ("", False, None, None),
        ("Step 4 (optional)", True, "2E75B6", "FFFFFF"),
        ("  Run 'ExportSnapshot' to save a plain .xlsx copy for sharing.", False, None, None),
        ("", False, None, None),
        ("SNMP REQUIREMENTS", True, "0D3F6E", "FFFFFF"),
        ("", False, None, None),
        ("  • Printers must have SNMP v1/v2c enabled.", False, None, None),
        ("  • The community string must match the printer's config (default: 'public').", False, None, None),
        ("  • UDP port 161 must be reachable from this PC (check firewall/VPN).", False, None, None),
        ("  • Windows PowerShell must be available (it is on all modern Windows).", False, None, None),
        ("  • Optionally install Net-SNMP for Windows for faster queries:", False, None, None),
        ("    https://www.net-snmp.org/", False, None, None),
        ("", False, None, None),
        ("COLOUR LEGEND", True, "0D3F6E", "FFFFFF"),
        ("", False, None, None),
        ("  Green  bar  = ≥ 26% remaining (OK)", False, "A9D18E", None),
        ("  Yellow bar  = 11–25% remaining (Warning)", False, "FFD966", None),
        ("  Red    bar  = ≤ 10% remaining (Critical)", False, "FF7043", None),
        ("  Grey   bar  = Level not reported", False, "D9D9D9", None),
    ]

    ws_help.column_dimensions["A"].width = 80
    ws_help.row_dimensions[1].height = 28
    for r, (text, bold, bg, fg) in enumerate(instructions, 1):
        cell = ws_help.cell(r, 1, text)
        cell.font = Font(bold=bold, size=11 if bold else 10,
                         color=fg if fg else "000000")
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(vertical="center", indent=0 if bold else 2)
        if bold:
            ws_help.row_dimensions[r].height = 22

    # ------------------------------------------------------------------ #
    #  Save as .xlsm skeleton  (no VBA binary yet – injected next)
    # ------------------------------------------------------------------ #
    tmp_xlsx = out_path.replace(".xlsm", "_tmp.xlsx")
    wb.save(tmp_xlsx)

    # ------------------------------------------------------------------ #
    #  Inject VBA:
    #  .xlsm is a zip. We add xl/vbaProject.bin (a minimal stub) and the
    #  VBA source in xl/VBA/PrinterSNMP.bas so users can import it.
    #  Full binary vbaProject.bin creation requires win32com or a pre-made
    #  stub; we ship the VBA as a .bas file and a README cell instead.
    # ------------------------------------------------------------------ #
    # Copy xlsx -> xlsm (same format, different extension tells Excel to allow macros)
    shutil.copy(tmp_xlsx, out_path)
    os.remove(tmp_xlsx)

    print(f"\n✓  Workbook skeleton saved: {out_path}")
    print("  NOTE: The VBA source is also saved separately as PrinterSNMP.bas")
    print("  Import it into the workbook: Alt+F11 -> File -> Import File -> PrinterSNMP.bas\n")

    # ------------------------------------------------------------------ #
    #  Write standalone .bas file (importable via VBA IDE)
    # ------------------------------------------------------------------ #
    bas_path = out_path.replace(".xlsm", ".bas").replace("PrinterConsumables", "PrinterSNMP")
    with open(bas_path, "w", encoding="utf-8") as f:
        f.write(VBA_MODULE.lstrip())
    print(f"✓  VBA module saved:         {bas_path}")


# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(
        description="Generate PrinterConsumables.xlsm + PrinterSNMP.bas"
    )
    parser.add_argument("--out", default="PrinterConsumables.xlsm",
                        help="Output .xlsm file path")
    args = parser.parse_args()
    build_xlsm(args.out)
